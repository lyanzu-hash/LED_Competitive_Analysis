"""
报告生成模块
Excel 结构（仅 2 个 Sheet）：
  Sheet1 - 今日竞品分析  （每行一个竞品，六维度 + 综合分析，彩色区分有无变化）
  Sheet2 - 竞品日报摘要  （LLM 综合日报全文）
"""

import os
import re
import logging
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import OUTPUT_DIR

logger = logging.getLogger(__name__)

# ── 颜色 ──────────────────────────────────────────────────────────────────────
C_HEADER   = "1F4E79"   # 深蓝（表头背景）
C_WHITE    = "FFFFFF"   # 白字
C_CHANGED  = "FFF2CC"   # 淡黄 - 今日有变化
C_NOCHANGE = "F2F2F2"   # 浅灰 - 今日无变化
C_FIRST    = "E2EFDA"   # 浅绿 - 首次建档
C_ACCENT   = "ED7D31"   # 橙色（大标题）
C_SECTION  = "D6E4F0"   # 节标题背景

# 六维度标签（与提示词保持一致）
DIMENSIONS = [
    "一、产品页面",
    "二、内容页面",
    "三、营销页面",
    "四、技术页面",
    "五、信任页面",
    "六、SEO维度",
    "七、综合分析",
]

_DIM_KEYS = ["product", "content", "marketing", "technical", "trust", "seo"]

_PAGE_TYPE_TO_DIM = {
    "产品页": "product",
    "博客/新闻": "content",
    "案例页": "content",
    "解决方案页": "marketing",
    "询盘/联系页": "marketing",
    "下载页": "technical",
    "FAQ页": "content",
    "关于页": "trust",
    "首页": "seo",
    "其他": "seo",
}

_FIELD_LABEL_TO_DIM = {
    # SEO
    "页面标题(Title)": "seo",
    "Meta描述": "seo",
    "H1标题": "seo",
    "H2标题(前8个)": "seo",
    "H3标题(前8个)": "seo",
    "Canonical URL": "seo",
    "多语言版本(hreflang)": "seo",
    # 产品/技术
    "价格信息": "product",
    "规格参数表": "product",
    "下载文件(Catalog/PDF)": "technical",
    # 营销
    "CTA行动按钮": "marketing",
    "表单入口": "marketing",
    # 信任
    "认证证书": "trust",
    "展会信息": "trust",
    # 内容
    "文章/博客标题": "content",
    # 技术/结构
    "视频嵌入": "technical",
    "Schema类型": "technical",
    "语言切换选项": "technical",
}


def _border() -> Border:
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def _hcell(ws, row, col, value, bg=C_HEADER, fg=C_WHITE, size=10, bold=True):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(bold=bold, color=fg, size=size)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = _border()
    return c


def _dcell(ws, row, col, value, bg=C_WHITE, size=9, bold=False):
    c = ws.cell(row=row, column=col, value=str(value) if value is not None else "")
    c.font      = Font(size=size, bold=bold)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(vertical="top", wrap_text=True)
    c.border    = _border()
    return c


def _set_width(ws, col, w):
    ws.column_dimensions[get_column_letter(col)].width = w


# ── 把 LLM 文本按 ## 标题拆分为字典 ─────────────────────────────────────────────
def _parse_sections(text: str) -> dict[str, str]:
    """
    将 LLM 输出的 markdown 文本按 '## 一、' 这类二级标题拆分。
    返回 { "一、产品页面": "内容...", ... }
    """
    sections: dict[str, str] = {}
    current_key = None
    buf: list[str] = []

    for line in text.splitlines():
        m = re.match(r"^##\s+(.+)", line.strip())
        if m:
            if current_key is not None:
                sections[current_key] = "\n".join(buf).strip()
            current_key = m.group(1).strip()
            buf = []
        else:
            buf.append(line)

    if current_key is not None:
        sections[current_key] = "\n".join(buf).strip()

    return sections


def _get_dim(sections: dict, dim_label: str) -> str:
    """模糊匹配维度标题，提取对应内容。"""
    key_num = dim_label.split("、")[0]   # "一" / "二" / ...
    for k, v in sections.items():
        if key_num in k:
            return v or "—"
    return "—"


def _guess_dim_from_url(url: str) -> str:
    u = (url or "").lower()
    if any(k in u for k in ["product", "led-display", "led-screen", "indoor", "outdoor", "rental", "cob", "gob"]):
        return "product"
    if any(k in u for k in ["blog", "news", "article", "post", "case", "project", "portfolio"]):
        return "content"
    if any(k in u for k in ["solution", "application", "industry", "landing", "campaign", "event"]):
        return "marketing"
    if any(k in u for k in ["download", "catalog", "datasheet", "manual", "resource"]):
        return "technical"
    if any(k in u for k in ["about", "company", "factory", "team", "certificate", "cert", "exhibition", "expo"]):
        return "trust"
    if any(k in u for k in ["contact", "inquiry", "quote", "rfq", "support"]):
        return "marketing"
    if any(k in u for k in ["faq", "help"]):
        return "content"
    return "seo"


def _classify_page_dim(url: str, page_type: str | None) -> str:
    pt = (page_type or "").strip()
    if pt and pt in _PAGE_TYPE_TO_DIM:
        return _PAGE_TYPE_TO_DIM[pt]
    return _guess_dim_from_url(url)


def _classify_field_dim(field_label: str) -> str:
    if field_label in _FIELD_LABEL_TO_DIM:
        return _FIELD_LABEL_TO_DIM[field_label]
    # 兜底：按关键词判断
    s = field_label.lower()
    if any(k in s for k in ["title", "meta", "h1", "h2", "h3", "canonical", "hreflang"]):
        return "seo"
    if any(k in s for k in ["price", "spec", "parameter"]):
        return "product"
    if any(k in s for k in ["cta", "form", "inquiry", "quote"]):
        return "marketing"
    if any(k in s for k in ["download", "video", "schema", "language"]):
        return "technical"
    if any(k in s for k in ["cert", "expo"]):
        return "trust"
    if any(k in s for k in ["blog", "article", "news"]):
        return "content"
    return "seo"


def _fmt_page_brief(title: str, h1_list: list | None, meta: str) -> str:
    parts = []
    if title:
        parts.append(f"Title={title[:120]}")
    if h1_list:
        h1 = " | ".join([str(x) for x in (h1_list or [])][:3]).strip()
        if h1:
            parts.append(f"H1={h1[:120]}")
    if meta:
        parts.append(f"Meta={meta[:160]}")
    return "；".join(parts) if parts else ""


def _build_dimension_cells(diff: dict | None) -> dict[str, str]:
    """
    将 diff 分发到六维度列，并按：
      变化信息 / 具体变化明细 / 前后对比
    组织文本。无变化则输出“与上次内容一致”。
    """
    if diff is None:
        return {k: "未启用对比（全量模式或无数据）" for k in _DIM_KEYS}
    if diff.get("is_first_run"):
        return {k: "首次建档，无对比基准" for k in _DIM_KEYS}
    if not diff.get("has_changes"):
        return {k: "与上次内容一致" for k in _DIM_KEYS}

    buckets: dict[str, dict] = {
        k: {"new": [], "removed": [], "changed": {}} for k in _DIM_KEYS
    }

    for it in diff.get("new_pages_detail", []) or []:
        url = it.get("url", "")
        dim = _classify_page_dim(url, it.get("page_type"))
        buckets[dim]["new"].append(it)

    for it in diff.get("removed_pages_detail", []) or []:
        url = it.get("url", "")
        dim = _classify_page_dim(url, it.get("page_type"))
        buckets[dim]["removed"].append(it)

    for item in diff.get("changed_pages", []) or []:
        url = item.get("url", "")
        for field_label, vals in (item.get("changes") or {}).items():
            dim = _classify_field_dim(field_label)
            by_url = buckets[dim]["changed"].setdefault(url, [])
            by_url.append({
                "field": field_label,
                "old": (vals.get("old") or "")[:300],
                "new": (vals.get("new") or "")[:300],
            })

    def _build_cell(bucket: dict) -> str:
        new_list = bucket["new"]
        rm_list = bucket["removed"]
        ch_map = bucket["changed"]
        if not new_list and not rm_list and not ch_map:
            return "与上次内容一致"

        lines: list[str] = []

        # 变化信息
        lines.append("变化信息：")
        lines.append(f"- 新增页面：{len(new_list)}")
        lines.append(f"- 下线页面：{len(rm_list)}")
        lines.append(f"- 内容变化页面：{len(ch_map)}")

        # 具体变化明细（不含前后值）
        lines.append("\n具体变化明细：")
        if new_list:
            lines.append("【新增】")
            for it in new_list:
                brief = _fmt_page_brief(it.get("title", ""), it.get("h1"), it.get("meta_description", ""))
                lines.append(f"+ {it.get('url','')}" + (f"\n  - {brief}" if brief else ""))
        if rm_list:
            lines.append("【下线】")
            for it in rm_list:
                brief = _fmt_page_brief(it.get("title", ""), it.get("h1"), it.get("meta_description", ""))
                lines.append(f"- {it.get('url','')}" + (f"\n  - {brief}" if brief else ""))
        if ch_map:
            lines.append("【字段变化】")
            for url, fields in ch_map.items():
                labels = "、".join([f["field"] for f in fields][:8])
                lines.append(f"~ {url}\n  - {labels}")

        # 前后对比（含 old/new）
        lines.append("\n前后对比：")
        if new_list:
            lines.append("【新增页面】")
            for it in new_list:
                brief = _fmt_page_brief(it.get("title", ""), it.get("h1"), it.get("meta_description", ""))
                lines.append(f"* {it.get('url','')}\n  - 旧：无\n  - 新：{brief or '（抓取到页面，但无可展示摘要）'}")
        if rm_list:
            lines.append("【下线页面】")
            for it in rm_list:
                brief = _fmt_page_brief(it.get("title", ""), it.get("h1"), it.get("meta_description", ""))
                lines.append(f"* {it.get('url','')}\n  - 旧：{brief or '（上次快照无摘要）'}\n  - 新：无")
        if ch_map:
            lines.append("【字段变化】")
            for url, fields in ch_map.items():
                lines.append(f"* {url}")
                for f in fields[:12]:
                    old = f.get("old") or "（空）"
                    new = f.get("new") or "（空）"
                    lines.append(f"  - {f['field']}\n    旧：{old}\n    新：{new}")

        return "\n".join(lines)

    return {k: _build_cell(buckets[k]) for k in _DIM_KEYS}


def _format_status_detail(diff: dict | None) -> str:
    """生成带数量统计的状态文本。"""
    if diff is None:
        return "🔵 未对比"
    if diff.get("is_first_run"):
        return "🟢 首次建档"
    if not diff.get("has_changes"):
        return "⚪ 今日无变化"
    parts = ["🟡 今日有更新"]
    counts = []
    if diff.get("new_pages"):
        counts.append(f"新增{len(diff['new_pages'])}页")
    if diff.get("removed_pages"):
        counts.append(f"下线{len(diff['removed_pages'])}页")
    if diff.get("changed_pages"):
        counts.append(f"{len(diff['changed_pages'])}页内容变化")
    if counts:
        parts.append("(" + "、".join(counts) + ")")
    return "\n".join(parts)


# ── Sheet 1：今日竞品分析 ─────────────────────────────────────────────────────
def _write_analysis_sheet(wb: openpyxl.Workbook, analyses: list[dict], diffs: dict, date_str: str):
    ws = wb.active
    ws.title = "今日竞品分析"

    col_count = 9  # 竞品 + 状态 + 六维度 + 综合分析
    ws.merge_cells(f"A1:{get_column_letter(col_count)}1")
    c = ws["A1"]
    c.value     = f"LED显示屏竞品日报  {date_str}  （由 AI 自动生成）"
    c.font      = Font(bold=True, size=14, color=C_ACCENT)
    c.fill      = PatternFill("solid", fgColor="FFF2CC")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    headers = ["竞品名称", "今日状态",
               "①产品页面", "②内容页面", "③营销页面",
               "④技术页面", "⑤信任页面", "⑥SEO维度",
               "综合分析（主推/关键词/风格/机会点）"]
    widths  = [14, 16, 44, 44, 44, 44, 44, 44, 55]

    for col, (h, w) in enumerate(zip(headers, widths), start=1):
        _hcell(ws, 2, col, h)
        _set_width(ws, col, w)
    ws.row_dimensions[2].height = 24

    for r, a in enumerate(analyses, start=3):
        name  = a.get("name", "")
        diff  = diffs.get(name)
        text  = a.get("analysis", "") or ""

        if diff is None or (isinstance(diff, dict) and diff.get("is_first_run")):
            bg = C_FIRST
        elif isinstance(diff, dict) and diff.get("has_changes"):
            bg = C_CHANGED
        else:
            bg = C_NOCHANGE

        status_text = _format_status_detail(diff)
        dim_cells = _build_dimension_cells(diff)

        _dcell(ws, r, 1, name, bg=bg, bold=True)
        _dcell(ws, r, 2, status_text, bg=bg)

        # 六维度列写入差异（按固定三段结构）
        col3_values = [
            dim_cells["product"],
            dim_cells["content"],
            dim_cells["marketing"],
            dim_cells["technical"],
            dim_cells["trust"],
            dim_cells["seo"],
        ]
        for col, content in enumerate(col3_values, start=3):
            _dcell(ws, r, col, content, bg=bg)

        # 综合分析：尽量抽取 LLM 的“七、综合分析”，若无则给占位
        summary_text = "—"
        if text and not text.startswith("（今日该竞品"):
            sections = _parse_sections(text)
            summary_text = _get_dim(sections, "七、综合分析")
        _dcell(ws, r, 9, summary_text, bg=bg)

        max_len = max([len(x) for x in col3_values + [summary_text]] + [0])
        ws.row_dimensions[r].height = min(520, max(80, max_len // 4))

    ws.freeze_panes = "A3"


# ── Sheet 2：竞品日报摘要 ─────────────────────────────────────────────────────
def _write_summary_sheet(wb: openpyxl.Workbook, date_str: str, summary_text: str):
    ws = wb.create_sheet("竞品日报摘要")

    ws.merge_cells("A1:B1")
    c = ws["A1"]
    c.value     = f"LED显示屏竞品日报  {date_str}"
    c.font      = Font(bold=True, size=15, color=C_ACCENT)
    c.fill      = PatternFill("solid", fgColor="FFF2CC")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:B2")
    ws["A2"].value     = "由 AI 大模型自动生成 · 每日更新"
    ws["A2"].font      = Font(italic=True, size=9, color="808080")
    ws["A2"].alignment = Alignment(horizontal="center")

    for i, line in enumerate(summary_text.splitlines(), start=3):
        ws.merge_cells(f"A{i}:B{i}")
        c = ws[f"A{i}"]
        c.value     = line
        c.alignment = Alignment(vertical="top", wrap_text=True)
        if line.startswith("## "):
            c.font = Font(bold=True, size=12, color="1F4E79")
            c.fill = PatternFill("solid", fgColor=C_SECTION)
        elif line.startswith("- ") or line.startswith("  -"):
            c.font = Font(size=10)
        else:
            c.font = Font(size=10)

    _set_width(ws, 1, 80)
    _set_width(ws, 2, 10)
    ws.freeze_panes = "A3"


# ── 主入口 ────────────────────────────────────────────────────────────────────
def save_report(
    scrape_results: list[dict],
    analyses:       list[dict],
    summary:        str,
    diffs:          dict | None = None,
) -> str:
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    date_str = datetime.now().strftime("%Y-%m-%d_%H%M")
    filepath = os.path.join(OUTPUT_DIR, f"LED竞品日报_{date_str}.xlsx")

    wb = openpyxl.Workbook()
    _write_analysis_sheet(wb, analyses, diffs or {}, date_str)
    _write_summary_sheet(wb, date_str, summary)

    wb.save(filepath)
    logger.info(f"[报告已保存] {filepath}")
    return filepath
